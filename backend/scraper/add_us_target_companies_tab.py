"""
Add a "US Target companies" tab to the RevEngineer workbook by:
1) Scanning job descriptions in the "AI-Filter-Titles" sheet for signals that
   the role targets the US market / US customers.
2) Writing a de-duplicated shortlist of companies into a new sheet.

Usage:
  py -3 add_us_target_companies_tab.py
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

IN_PATH = "results/RevEngineer_India_Marketing_LAST30D.xlsx"
IN_SHEET = "AI-Filter-Titles"
OUT_SHEET = "US Target companies"


@dataclass(frozen=True)
class Match:
    name: str
    weight: int
    snippet: str


def _strip_illegal_excel_chars(value: str) -> str:
    # Strip control chars openpyxl/Excel can't store.
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value or "")


def _normalize(text: str) -> str:
    t = (text or "").replace("\u2011", "-")
    t = _strip_illegal_excel_chars(t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _compile_patterns() -> tuple[list[tuple[str, re.Pattern, int]], list[tuple[str, re.Pattern, int]]]:
    # IMPORTANT: do NOT let "US" match the pronoun "us".
    # We keep patterns case-insensitive overall, but force US/USA to be uppercase.
    # "united states" can be any case.
    us_token = r"(?:\b(?-i:US)\b|\b(?-i:USA)\b|\bU\.S\.A\.\b|\bU\.S\.\b|united\s+states)"
    flags = re.I | re.S

    pos: list[tuple[str, re.Pattern, int]] = [
        ("US market focus", re.compile(rf"{us_token}\W+market\W+focus\b", flags), 6),
        ("Expand into US", re.compile(rf"\bexpand(?:ing)?\b(?:\W+\w+){{0,10}}\W+(?:into|in|to)\W+(?:the\W+)?{us_token}\b", flags), 6),
        ("Enter US market", re.compile(rf"\benter(?:ing)?\b(?:\W+\w+){{0,12}}\W+{us_token}(?:\W+\w+){{0,6}}\W+markets?\b", flags), 6),
        ("Build US presence", re.compile(rf"\bbuild(?:ing)?\b(?:\W+\w+){{0,12}}\W+(?:a\W+)?{us_token}(?:\W+\w+){{0,6}}\W+(presence|business|pipeline|customer\s+base)\b", flags), 6),
        ("Establish US office/team", re.compile(rf"\b(?:establish|opening|open|set\W*up|build)\b(?:\W+\w+){{0,12}}\W+(?:a\W+)?{us_token}(?:\W+\w+){{0,6}}\W+(office|team|operations|subsidiary)\b", flags), 6),

        ("US market(s)", re.compile(rf"{us_token}\W+markets?\b", flags), 4),
        ("US healthcare market", re.compile(rf"{us_token}\W+healthcare\W+market\b", flags), 4),
        ("US customers/clients", re.compile(rf"{us_token}\W+(customers|clients|accounts|logos|buyers|prospects)\b", flags), 4),
        ("Targeting US", re.compile(rf"\btarget(?:ing)?\b(?:\W+\w+){{0,10}}\W+{us_token}\b", flags), 4),
        ("Sell into US", re.compile(rf"\bsell(?:ing)?\b(?:\W+\w+){{0,10}}\W+into\W+(?:the\W+)?{us_token}\b", flags), 4),
        ("US demand gen/pipeline", re.compile(rf"{us_token}\W+(pipeline|demand\W+gen|demand\W+generation|lead\W+generation|outbound|territory)\b", flags), 4),

        ("US stakeholders/teams", re.compile(rf"{us_token}\W+(stakeholders?|leadership|sales|teams?)\b", flags), 3),
        ("Markets list includes US", re.compile(rf"\bmarkets?\b[^\n]{{0,120}}{us_token}\b", flags), 3),
        ("Global campaigns incl. US", re.compile(rf"\b(global|international)\b(?:\W+\w+){{0,4}}\W+campaigns?\b(?:\W+\w+){{0,12}}\W+{us_token}\b", flags), 3),
        ("Across India/US/EU markets", re.compile(r"\bacross\b[^\n]{0,120}\bindia\b[^\n]{0,120}\b(us|u\.s\.|usa|united states)\b[^\n]{0,120}\b(eu|europe)\b", flags), 3),

        ("US buying landscape", re.compile(rf"{us_token}\W+(technology\W+)?buying\W+landscape\b", flags), 2),
        ("North America market/customers", re.compile(r"\b(north america|north american)\b\W+(market|markets|customers|clients)\b", flags), 3),
    ]

    neg: list[tuple[str, re.Pattern, int]] = [
        ("US work authorization / visa", re.compile(rf"{us_token}(?:\W+\w+){{0,10}}\W+(work\W+authorization|authorized\W+to\W+work|\bvisa\b|\bh1b\b|\bopt\b|e-verify|\bw2\b|\b1099\b)\b", flags), -5),
        ("US remote/location only", re.compile(rf"{us_token}\W*(remote|based)\b|\bmust\W+be\W+located\W+in\W+(?:the\W+)?{us_token}\b", flags), -4),
        ("US English only", re.compile(r"\bUS\W+English\b", flags), -3),
        ("Comp/location boilerplate", re.compile(r"\b(employees?\s+dispersed\s+throughout\s+the\s+united\s+states|market-based\s+pay\s+structure)\b", flags), -3),
    ]

    return pos, neg


def score_row(title: str, description: str) -> tuple[int, str, list[Match]]:
    pos, neg = _compile_patterns()

    t = _normalize(f"{title or ''}\n{description or ''}")
    score = 0
    matches: list[Match] = []

    for name, pat, weight in pos:
        m = pat.search(t)
        if not m:
            continue
        score += weight
        s, e = m.span()
        snippet = t[max(0, s - 90) : min(len(t), e + 140)]
        matches.append(Match(name=name, weight=weight, snippet=snippet))

    for _name, pat, weight in neg:
        if pat.search(t):
            score += weight

    # Require at least one core US-market signal.
    core = {
        "US market focus",
        "Expand into US",
        "Enter US market",
        "Build US presence",
        "Establish US office/team",
        "US market(s)",
        "US healthcare market",
        "US customers/clients",
        "Targeting US",
        "Sell into US",
        "US demand gen/pipeline",
        "US stakeholders/teams",
        "Markets list includes US",
        "Global campaigns incl. US",
        "Across India/US/EU markets",
        "North America market/customers",
    }
    has_core = any(m.name in core for m in matches)
    if not has_core:
        return score, "Unlikely", matches

    # Classify
    if any(m.weight >= 6 for m in matches) or score >= 8:
        cls = "Likely"
    elif score >= 3:
        cls = "Possible"
    else:
        cls = "Unlikely"
    return score, cls, matches


def _best_match_snippet(matches: Iterable[Match]) -> str:
    ms = list(matches)
    if not ms:
        return ""
    ms.sort(key=lambda m: (m.weight, len(m.snippet)), reverse=True)
    # Keep snippets compact for Excel readability.
    return ms[0].snippet[:600]


def main() -> None:
    df = pd.read_excel(IN_PATH, sheet_name=IN_SHEET, dtype=str).fillna("")

    scored_rows = []
    for _, r in df.iterrows():
        score, cls, matches = score_row(r.get("title", ""), r.get("description", ""))
        if cls == "Unlikely":
            continue
        scored_rows.append(
            {
                "companyName": r.get("companyName", "").strip(),
                "classification": cls,
                "score": score,
                "evidence": _best_match_snippet(matches),
                "matchedSignals": "; ".join(sorted({m.name for m in matches})),
                "sampleTitle": r.get("title", ""),
                "sampleLocation": r.get("location", ""),
                "sampleDatePosted": r.get("datePosted", ""),
                "sampleSite": r.get("site", ""),
                "sampleJobUrl": r.get("jobUrl", ""),
            }
        )

    if not scored_rows:
        raise SystemExit("No US-targeting companies detected in AI-Filter-Titles.")

    scored = pd.DataFrame(scored_rows)

    # De-duplicate by company, keeping the strongest evidence row.
    scored = scored.sort_values(["classification", "score"], ascending=[True, False])
    scored = scored.sort_values(["score"], ascending=[False]).reset_index(drop=True)
    scored["postsMatched"] = scored.groupby("companyName")["companyName"].transform("count")
    best = scored.sort_values(["score"], ascending=[False]).drop_duplicates("companyName", keep="first")

    # Web evidence (lightweight manual mapping from our targeted web lookups).
    web_info = {
        "2Base Technologies": {
            "webEvidence": "Company materials list registered offices in the USA and UK (in addition to India), supporting US/UK market delivery.",
            "webSources": "https://www.2basetechnologies.com/; https://www.linkedin.com/company/2base-technologies",
        },
        "Aurigo Software Technologies": {
            "webEvidence": "Public company profile describes Aurigo as a U.S. corporation headquartered in Austin, Texas, with projects/customers across North America.",
            "webSources": "https://www.linkedin.com/company/aurigo-software-technologies",
        },
        "Comviva": {
            "webEvidence": "Public materials describe Comviva as a global digital solutions provider (subsidiary of Tech Mahindra).",
            "webSources": "https://www.comviva.com/corporate/contact-us/; https://en.wikipedia.org/wiki/Comviva",
        },
        "eG Innovations": {
            "webEvidence": "Company describes itself as providing IT monitoring/observability solutions with global presence (including US).",
            "webSources": "https://www.eginnovations.com/company/about-us",
        },
        "Kavi India": {
            "webEvidence": "Company profile lists headquarters in Chennai and describes the organization and its business scope on LinkedIn.",
            "webSources": "https://www.linkedin.com/company/kavi-india",
        },
        "Kovalent Coatings International": {
            "webEvidence": "Company website positions Kovalent Coatings as a global ceramic coating manufacturer; job text is the primary US-market signal here.",
            "webSources": "https://www.kovalentcoatings.com/",
        },
        "MTechZilla": {
            "webEvidence": "Company site positions itself as an offshore engineering partner delivering to companies across the US and Europe.",
            "webSources": "https://www.mtechzilla.com/",
        },
        "Narwal": {
            "webEvidence": "Public profiles list Narwal as headquartered in Cincinnati, Ohio (US) with presence across North America/UK/Mexico/India.",
            "webSources": "https://www.linkedin.com/company/narwal-ai; https://craft.co/narwal/locations",
        },
        "Pepper": {
            "webEvidence": "Company materials and job listings describe a US-market demand-gen motion (e.g., 'Head of Demand Generation (US Market)').",
            "webSources": "https://www.pepper.inc/pepper-ai; https://in.linkedin.com/company/pepperinc/jobs",
        },
        "Prolim": {
            "webEvidence": "Public materials describe PROLIM as headquartered in the US (Farmington Hills, MI) with a multi-country office footprint.",
            "webSources": "https://www.michiganbusiness.org/4af065/globalassets/documents/international-trade-services/prolim-success-story.pdf; https://www.einpresswire.com/article/822017628/prolim-celebrates-20-years-of-global-innovation-and-excellence",
        },
        "Recro": {
            "webEvidence": "Recro positions itself as a platform and services provider; job text itself is the primary US-market targeting signal here.",
            "webSources": "https://www.recro.io/",
        },
        "SourcingXPress": {
            "webEvidence": "Job listing explicitly states 'US Market Focus' (CMO role).",
            "webSources": "https://in.linkedin.com/jobs/view/chief-marketing-officer-at-sourcingxpress-4411602827; https://www.sourcingxpress.com/about",
        },
        "Unosecur": {
            "webEvidence": "Company website positions Unosecur as an identity security platform; job text is the primary US-market signal here.",
            "webSources": "https://www.unosecur.com/; https://www.unosecur.com/about-us",
        },
        "Uplers": {
            "webEvidence": "Public materials describe Uplers as a hiring platform serving global companies with remote talent from India.",
            "webSources": "https://www.prnewswire.com/news-releases/uplers-is-shortening-hiring-cycle-from-4-months-to-48-hours-for-global-digital-agencies-with-deeply-vetted-indian-talents-301984668.html",
        },
        "Lytegen": {
            "webEvidence": "Company site describes operating across multiple US states; the job post is explicitly labeled 'US Market'.",
            "webSources": "https://lytegen.com/; https://www.marketingmonk.so/jobboard/jobs/head-of-performance-marketing-b2c-us-market-at-lytegen-3h7noa",
        },
        "Varun Digital Media": {
            "webEvidence": "Company website markets itself as serving clients across the USA and lists areas served in the USA.",
            "webSources": "https://www.varundigitalmedia.com/; https://www.varundigitalmedia.com/area-we-serve",
        },
    }

    best["webEvidence"] = best["companyName"].map(lambda c: web_info.get(c, {}).get("webEvidence", ""))
    best["webSources"] = best["companyName"].map(lambda c: web_info.get(c, {}).get("webSources", ""))

    # Order columns for the output sheet.
    out_cols = [
        "companyName",
        "classification",
        "score",
        "postsMatched",
        "matchedSignals",
        "evidence",
        "sampleTitle",
        "sampleLocation",
        "sampleDatePosted",
        "sampleSite",
        "sampleJobUrl",
        "webEvidence",
        "webSources",
    ]
    best = best[out_cols].sort_values(["classification", "score"], ascending=[True, False]).reset_index(drop=True)

    wb = load_workbook(IN_PATH)
    if OUT_SHEET in wb.sheetnames:
        del wb[OUT_SHEET]
    ws = wb.create_sheet(OUT_SHEET, 0)

    # Header styling consistent with build_excel.py
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    headers = list(best.columns)
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center")

    # Write rows
    for _, row in best.iterrows():
        ws.append([_strip_illegal_excel_chars(str(row.get(h, ""))) for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Formatting
    wrap_cols = {"evidence", "webEvidence", "webSources"}
    widths = {
        "companyName": 28,
        "classification": 12,
        "score": 8,
        "postsMatched": 12,
        "matchedSignals": 34,
        "evidence": 80,
        "sampleTitle": 42,
        "sampleLocation": 26,
        "sampleDatePosted": 12,
        "sampleSite": 14,
        "sampleJobUrl": 48,
        "webEvidence": 56,
        "webSources": 70,
    }
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 14)
        if h in wrap_cols:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(IN_PATH)
    print(f"Added sheet: {OUT_SHEET} ({len(best)} companies) -> {IN_PATH}")


if __name__ == "__main__":
    main()
