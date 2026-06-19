"""Build the RevEngineer India-marketing workbook:
  Tab 1 "ALL jobs Posts"   -> the full last-30-days CSV
  Tab 2 "AI-Filter-Titles" -> posts whose title is a good target for revengineer.ai
                              (revenue / growth / demand-gen / RevOps / perf /
                               product marketing + senior marketing leadership),
                              with a RevEngineerFit category column.
"""
import re
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

SRC = "results/jobs_india_marketing_LAST30D_full_2026-05-25.csv"
OUT = "results/RevEngineer_India_Marketing_LAST30D.xlsx"

# Narrow (lower-fit) marketing specialisms for a revenue-AI tool.
NARROW = re.compile(r"\b(brand|content|social|influencer|trade|communications?|public relations|\bpr\b|creative|design|event|field marketing|community)\b", re.I)


def fit_category(title):
    t = (title or "").lower()
    if "cmo" in re.sub(r"[^a-z ]", " ", t).split() or "chief marketing" in t:
        return "CMO / Chief Marketing"
    if re.search(r"demand gen", t):
        return "Demand Generation"
    if "revenue marketing" in t or re.search(r"\brevops\b|revenue operations", t):
        return "Revenue Marketing / RevOps"
    if re.search(r"\bgrowth\b", t):
        return "Growth"
    if re.search(r"marketing op|marketing operations|martech", t):
        return "Marketing Ops"
    if "performance marketing" in t or "demand generation" in t:
        return "Performance / Demand Gen"
    if "product marketing" in t:
        return "Product Marketing"
    if re.search(r"digital marketing", t):
        return "Digital Marketing"
    # General senior marketing leadership (owns revenue) — but skip if it's a
    # narrow brand/content/social/trade/comms leader.
    if re.search(r"\b(head|chief|vp|vice president|director|general manager|gm)\b", t) and "marketing" in t:
        return "" if NARROW.search(t) else "Marketing Leadership"
    if re.search(r"marketing\s*(head|director|lead)", t):
        return "" if NARROW.search(t) else "Marketing Leadership"
    return ""


def main():
    df = pd.read_csv(SRC, dtype=str).fillna("")
    # Strip control chars openpyxl/Excel can't store.
    illegal = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
    clean = lambda v: illegal.sub("", v) if isinstance(v, str) else v
    df = df.map(clean) if hasattr(df, "map") else df.applymap(clean)
    cat = df["title"].map(fit_category)
    flt = df[cat != ""].copy()
    flt.insert(2, "RevEngineerFit", cat[cat != ""].values)
    # sort filtered by fit category then company
    flt = flt.sort_values(["RevEngineerFit", "companyName"]).reset_index(drop=True)

    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="ALL jobs Posts", index=False)
        flt.to_excel(xw, sheet_name="AI-Filter-Titles", index=False)

        wb = xw.book
        for name in ("ALL jobs Posts", "AI-Filter-Titles"):
            ws = wb[name]
            ws.freeze_panes = "A2"
            hdr_fill = PatternFill("solid", fgColor="1F4E78")
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = hdr_fill
                cell.alignment = Alignment(vertical="center")
            # column widths
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            widths = {"title": 42, "RevEngineerFit": 24, "companyName": 28, "location": 26,
                      "jobUrl": 46, "datePosted": 12, "description": 80, "site": 14,
                      "id": 16, "jobType": 14}
            for c, h in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 12)
                if h == "description":
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.auto_filter.ref = ws.dimensions

    print(f"ALL jobs Posts: {len(df)} rows")
    print(f"AI-Filter-Titles: {len(flt)} rows")
    print("fit breakdown:")
    for k, v in flt["RevEngineerFit"].value_counts().items():
        print(f"  {k}: {v}")
    print("saved:", OUT)


if __name__ == "__main__":
    main()
