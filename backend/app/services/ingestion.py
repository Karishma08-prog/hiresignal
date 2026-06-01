from __future__ import annotations

import csv
import hashlib
import re
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

from sqlalchemy.orm import Session
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

from app import models
from app.config import settings
from app.services.source_catalog import SOURCE_METADATA
from app.services.source_registry import record_source_evidence, refresh_source_support
from app.utils import make_id


@dataclass
class OutputFile:
    path: Path
    kind: str
    source_key: str | None = None


def _normalize_site_key(value: str | None) -> str:
    raw = (value or "").strip().lower()
    if not raw:
        return "unknown"
    if ":" in raw:
        head = raw.split(":", 1)[0].strip()
        if head in {
            "greenhouse",
            "lever",
            "workday",
            "ashby",
            "smartrecruiters",
            "jobvite",
            "workable",
            "rippling",
            "bamboohr",
            "personio",
            "jazzhr",
            "recruitee",
            "teamtailor",
            "icims",
            "taleo",
            "successfactors",
            "adp",
            "ukg",
            "breezyhr",
            "comeet",
            "pinpoint",
            "manatal",
            "paylocity",
            "freshteam",
            "bullhorn",
            "trakstar",
            "hiringthing",
            "loxo",
            "fountain",
            "deel",
            "phenom",
            "jobylon",
            "homerun",
            "jobscore",
            "talentlyft",
            "crelate",
            "ismartrecruit",
            "recruiterflow",
        }:
            return head
    if raw.startswith("workday:"):
        return "workday"
    return (
        raw.replace(" ", "_")
        .replace("-", "_")
        .replace("/", "_")
        .replace(".", "_")
        .replace(":", "_")
    )


def _title_case_site(site_key: str) -> str:
    parts = [part for part in site_key.replace("_", " ").split(" ") if part]
    return " ".join(part.capitalize() for part in parts) or "Unknown"


def ensure_source(db: Session, site_key: str, *, engine_hint: str | None = None) -> models.Source:
    normalized = _normalize_site_key(site_key)
    source = db.query(models.Source).filter(models.Source.site_key == normalized).first()
    if source is not None:
        return source

    meta = SOURCE_METADATA.get(normalized, {})
    source = models.Source(
        id=make_id("src"),
        site_key=normalized,
        display_name=str(meta.get("display_name", _title_case_site(normalized))),
        category=str(meta.get("category", "search_board")),
        engine=str(meta.get("engine", engine_hint or "script_bridge")),
        region=str(meta.get("region", "global")),
        requires_company_slug=bool(meta.get("requires_company_slug", False)),
        requires_api_key=bool(meta.get("requires_api_key", False)),
        risk_level=str(meta.get("risk_level", "secondary")),
        notes=str(meta.get("notes", "Imported from scraper output.")),
    )
    db.add(source)
    db.flush()
    return source


def ensure_source_health(db: Session, site_key: str) -> models.SourceHealth:
    normalized = _normalize_site_key(site_key)
    source = ensure_source(db, normalized)
    health = (
        db.query(models.SourceHealth).filter(models.SourceHealth.site_key == normalized).first()
    )
    if health is not None:
        return health
    health = models.SourceHealth(
        id=make_id("srchealth"),
        site_key=source.site_key,
        status="ready",
        avg_results_7d=0,
        avg_latency_ms_7d=0,
        success_rate_7d=0,
        last_run_jobs_found=0,
    )
    db.add(health)
    db.flush()
    return health


def ensure_source_credential(db: Session, site_key: str) -> models.SourceCredential:
    normalized = _normalize_site_key(site_key)
    source = ensure_source(db, normalized)
    credential = (
        db.query(models.SourceCredential)
        .filter(models.SourceCredential.site_key == normalized)
        .first()
    )
    if credential is not None:
        return credential

    meta = SOURCE_METADATA.get(normalized, {})
    needs_api_key = bool(meta.get("needs_api_key", False))
    needs_proxy = bool(meta.get("needs_proxy", False))
    needs_company_slug = bool(meta.get("requires_company_slug", False))
    credential_present = (
        (needs_api_key and bool(settings.scrappa_token))
        or (needs_proxy and bool(settings.jobs_bota_proxy or settings.jobs_proxy))
        or (not needs_api_key and not needs_proxy)
    )
    credential = models.SourceCredential(
        id=make_id("cred"),
        site_key=source.site_key,
        needs_api_key=needs_api_key,
        needs_proxy=needs_proxy,
        needs_company_slug=needs_company_slug,
        credential_present=credential_present,
        credential_verified_at=datetime.utcnow() if credential_present else None,
        working_status="unknown",
        credential_note=(
            "Uses SCRAPPA_TOKEN for ATS discovery."
            if needs_api_key
            else "Uses proxy-backed browser scraping." if needs_proxy else "No credential required."
        ),
    )
    db.add(credential)
    db.flush()
    return credential


def snapshot_results(results_dir: Path) -> dict[str, float]:
    if not results_dir.exists():
        return {}
    return {str(path): path.stat().st_mtime for path in results_dir.iterdir() if path.is_file()}


def discover_output_files(
    campaign: models.Campaign,
    before_snapshot: dict[str, float] | None = None,
    *,
    allow_fallback: bool = True,
    limit: int = 8,
) -> list[OutputFile]:
    results_dir = settings.results_dir
    if not results_dir.exists():
        return []

    before_snapshot = before_snapshot or {}
    changed: list[Path] = []
    for path in results_dir.iterdir():
        if not path.is_file():
            continue
        prev = before_snapshot.get(str(path))
        current = path.stat().st_mtime
        if prev is None or current > prev:
            changed.append(path)

    if changed:
        candidates = sorted(changed, key=lambda p: p.stat().st_mtime, reverse=True)
    else:
        if not allow_fallback:
            return []
        candidates = _campaign_fallback_candidates(campaign, results_dir)

    outputs: list[OutputFile] = []
    for path in candidates[:limit]:
        kind = path.suffix.lower().lstrip(".")
        source_key = None
        if path.name.startswith("ats_"):
            source_key = "ats_discovery"
        outputs.append(OutputFile(path=path, kind=kind, source_key=source_key))
    return outputs


def _campaign_fallback_candidates(campaign: models.Campaign, results_dir: Path) -> list[Path]:
    name = campaign.name.lower()
    files = [path for path in results_dir.iterdir() if path.is_file()]

    def select(patterns: list[str]) -> list[Path]:
        matched = []
        seen: set[Path] = set()
        for pattern in patterns:
            for path in sorted(results_dir.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True):
                if path not in seen:
                    seen.add(path)
                    matched.append(path)
        return matched

    if "ats" in name:
        chosen = select(["ats_*.csv", "*_ATS_*.csv", "*WORKDAY*.csv", "*.xlsx"])
    elif campaign.country.upper() == "INDIA" or "india" in name:
        chosen = select(
            [
                "jobs_india_marketing_*.csv",
                "jobs_india_marketing_LAST30D*.csv",
                "extra_boards_*.csv",
                "ats_marketing_*.csv",
                "*WORKDAY*.csv",
                "RevEngineer_*.xlsx",
            ]
        )
    else:
        chosen = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)
    return chosen or sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)


def _csv_has_job_schema(path: Path) -> bool:
    try:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            headers = next(reader, [])
    except Exception:
        return False
    normalized = {header.strip() for header in headers}
    required = {"id", "site", "title", "companyName", "jobUrl"}
    return required.issubset(normalized)


def _csv_has_ats_schema(path: Path) -> bool:
    try:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            headers = next(reader, [])
    except Exception:
        return False
    normalized = {header.strip() for header in headers}
    return {"ats", "companySlug", "jobUrl"}.issubset(normalized)


def _parse_bool(value: str | None) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y"}


def _parse_float(value: str | None) -> float | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return float(raw)
    except Exception:
        return None


def _safe_date(value: str | None) -> str | None:
    raw = (value or "").strip()
    if not raw:
        return None
    return raw[:10]


def _extract_domain(url: str | None) -> str | None:
    if not url:
        return None
    try:
        hostname = urlparse(url).hostname or ""
    except Exception:
        return None
    if hostname.startswith("www."):
        hostname = hostname[4:]
    return hostname or None


def _company_key(name: str) -> str:
    return " ".join(name.lower().split())


def _hash_job(site: str, job_url: str, title: str, company_name: str) -> str:
    digest = hashlib.sha1(f"{site}|{job_url}|{title}|{company_name}".encode("utf-8")).hexdigest()
    return digest[:24]


def _campaign_title_terms(campaign: models.Campaign) -> list[str]:
    title_config = campaign.title_filter_config or {}
    configured_titles = [
        str(item).strip().lower()
        for item in title_config.get("includeTitles", [])
        if str(item).strip()
    ]
    role_query = campaign.role_query or ""
    query_parts = [
        part.strip().lower()
        for part in re.split(r"\bor\b", role_query, flags=re.IGNORECASE)
        if part.strip()
    ]

    ordered_terms: list[str] = []
    seen: set[str] = set()
    for term in [*configured_titles, *query_parts]:
        if term and term not in seen:
            seen.add(term)
            ordered_terms.append(term)
    return ordered_terms


def _matched_title(job_title: str, campaign: models.Campaign) -> bool:
    title = (job_title or "").lower()
    terms = _campaign_title_terms(campaign)
    return any(term in title for term in terms) if terms else False


def _objective_signals(description: str, campaign: models.Campaign) -> list[str]:
    text = (description or "").lower()
    objective = campaign.objective_filter_config or {}
    configured = [
        str(item).strip()
        for item in objective.get("signals", [])
        if str(item).strip()
    ]
    signals = configured or [
        "us market",
        "united states",
        "north america",
        "us customers",
        "expand into us",
    ]
    return [signal for signal in signals if signal.lower() in text]


def _classify_fit(matched_title: bool, matched_objective: bool) -> tuple[str, str]:
    if matched_title and matched_objective:
        return "high", "high"
    if matched_title or matched_objective:
        return "medium", "medium"
    return "low", "low"


def _copy_artifact(path: Path, run_id: str) -> Path:
    normalized_name = path.stem.lower()
    if "company-shortlist" in normalized_name:
        label = "company-shortlist"
    elif "search" in normalized_name:
        label = "search"
    elif "browser" in normalized_name:
        label = "browser"
    elif "ats" in normalized_name:
        label = "ats"
    elif path.suffix.lower() == ".xlsx":
        label = "report"
    else:
        label = "artifact"

    fingerprint = hashlib.sha1(str(path).encode("utf-8")).hexdigest()[:12]
    destination = settings.artifacts_dir / f"{run_id}-{label}-{fingerprint}{path.suffix.lower()}"
    shutil.copy2(path, destination)
    return destination


def output_file_matches_campaign(path: Path, campaign: models.Campaign, *, sample_limit: int = 250) -> bool:
    if not _campaign_title_terms(campaign):
        return True

    suffix = path.suffix.lower()
    if suffix == ".csv" and (_csv_has_job_schema(path) or _csv_has_ats_schema(path)):
        try:
            with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
                reader = csv.DictReader(handle)
                for index, row in enumerate(reader):
                    if index >= sample_limit:
                        break
                    if _matched_title(str(row.get("title") or ""), campaign):
                        return True
        except Exception:
            return False
        return False

    if suffix == ".xlsx":
        rows = _load_us_target_shortlist(path)
        for row in rows[:sample_limit]:
            sample_title = str(row.get("sampleTitle") or row.get("titleMatch") or "")
            if _matched_title(sample_title, campaign):
                return True
        return False

    return False


def _load_us_target_shortlist(path: Path) -> list[dict[str, str]]:
    if load_workbook is None or path.suffix.lower() != ".xlsx":
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        if "US Target companies" not in workbook.sheetnames:
            return []
        sheet = workbook["US Target companies"]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        output: list[dict[str, str]] = []
        for row in rows[1:]:
            item = {
                headers[index]: str(value).strip()
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
            if item.get("companyName"):
                output.append(item)
        return output
    finally:
        workbook.close()


def _generate_shortlist_artifact(db: Session, company_ids: list[str], run_id: str) -> Path:
    path = settings.artifacts_dir / f"{run_id}-company-shortlist.csv"
    companies = (
        db.query(models.Company)
        .filter(models.Company.id.in_(company_ids))
        .order_by(models.Company.name.asc())
        .all()
    )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "company_name",
                "location",
                "objective_signal",
                "title_match",
                "open_roles",
                "priority",
                "fit",
                "source",
            ]
        )
        for company in companies:
            writer.writerow(
                [
                    company.name,
                    company.location or "",
                    company.objective_signal or "",
                    company.title_match or "",
                    company.open_roles,
                    company.priority,
                    company.revengineer_fit,
                    company.source or "",
                ]
            )
    return path


def ingest_campaign_outputs(
    db: Session,
    campaign: models.Campaign,
    run: models.CampaignRun,
    output_files: list[OutputFile],
    source_summary: list[dict],
    *,
    ingestion_mode: str = "historical_import",
    restrict_to_campaign_title_match: bool = False,
    mark_empty_results_as_failed: bool = False,
    allowed_source_keys: set[str] | None = None,
    source_evidence_type: str | None = None,
) -> dict[str, object]:
    job_rows: list[dict] = []
    ats_rows: list[dict] = []
    shortlist_rows: list[dict[str, str]] = []

    for output in output_files:
        if output.kind == "csv" and _csv_has_job_schema(output.path):
            with output.path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
                reader = csv.DictReader(handle)
                job_rows.extend(list(reader))
        elif output.kind == "csv" and _csv_has_ats_schema(output.path):
            with output.path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
                reader = csv.DictReader(handle)
                ats_rows.extend(list(reader))
        elif output.kind == "xlsx":
            shortlist_rows.extend(_load_us_target_shortlist(output.path))

    normalized_allowed_sources = {
        _normalize_site_key(item)
        for item in (allowed_source_keys or set())
        if str(item).strip()
    }
    if normalized_allowed_sources:
        job_rows = [
            row
            for row in job_rows
            if _normalize_site_key(str(row.get("site") or "")) in normalized_allowed_sources
        ]
        ats_rows = [
            row
            for row in ats_rows
            if _normalize_site_key(str(row.get("ats") or "")) in normalized_allowed_sources
        ]

    best_job_rows: dict[str, dict] = {}
    for raw in job_rows:
        site_key = _normalize_site_key(raw.get("site"))
        company_name = (raw.get("companyName") or "").strip() or "Unknown company"
        title = (raw.get("title") or "").strip() or "Untitled role"
        job_url = (raw.get("jobUrl") or "").strip()
        if not job_url:
            continue
        normalized_hash = _hash_job(site_key, job_url, title, company_name)
        existing = best_job_rows.get(normalized_hash)
        if existing is None:
            best_job_rows[normalized_hash] = raw
            continue
        existing_desc_len = len((existing.get("description") or "").strip())
        current_desc_len = len((raw.get("description") or "").strip())
        existing_site = _normalize_site_key(existing.get("site"))
        current_site = _normalize_site_key(raw.get("site"))
        if current_desc_len > existing_desc_len or (
            current_site == "naukri" and current_desc_len >= existing_desc_len
        ):
            best_job_rows[normalized_hash] = raw

    job_rows = list(best_job_rows.values())
    if restrict_to_campaign_title_match and _campaign_title_terms(campaign):
        job_rows = [row for row in job_rows if _matched_title(str(row.get("title") or ""), campaign)]

    companies_by_key: dict[str, models.Company] = {}
    company_metrics: dict[str, dict[str, object]] = defaultdict(
        lambda: {
            "jobs": 0,
            "locations": [],
            "sources": [],
            "descriptions": [],
            "days_active": [],
            "signals": [],
            "title_matches": [],
        }
    )
    created_jobs: list[models.Job] = []
    per_site_counts: Counter[str] = Counter()

    for raw in job_rows:
        site_key = _normalize_site_key(raw.get("site"))
        company_name = (raw.get("companyName") or "").strip() or "Unknown company"
        title = (raw.get("title") or "").strip() or "Untitled role"
        job_url = (raw.get("jobUrl") or "").strip()
        if not job_url:
            continue
        normalized_hash = _hash_job(site_key, job_url, title, company_name)

        source = ensure_source(db, site_key)
        ensure_source_credential(db, site_key)
        key = _company_key(company_name)
        company = companies_by_key.get(key)
        if company is None:
            company = (
                db.query(models.Company)
                .filter(models.Company.name.ilike(company_name))
                .first()
            )
            if company is None:
                company = models.Company(
                    id=make_id("com"),
                    name=company_name,
                    website=None,
                    domain=_extract_domain(job_url),
                    industry=None,
                    location=(raw.get("location") or "").strip() or None,
                    description=(raw.get("description") or "").strip() or None,
                    open_roles=0,
                    status="active",
                    revengineer_fit="medium",
                    priority="medium",
                    objective_signal=None,
                    title_match=None,
                    days_active=0,
                    source=source.display_name,
                    web_evidence=None,
                    web_sources=[],
                )
                db.add(company)
                db.flush()
            companies_by_key[key] = company

        matched_title = _matched_title(title, campaign)
        matched_signals = _objective_signals(raw.get("description") or "", campaign)
        matched_objective = bool(matched_signals)
        fit, priority = _classify_fit(matched_title, matched_objective)

        job = models.Job(
            id=make_id("job"),
            campaign_run_id=run.id,
            company_id=company.id,
            site=site_key,
            engine=source.engine,
            external_id=(raw.get("id") or "").strip() or None,
            job_url=job_url,
            title=title,
            company_name=company_name,
            location=(raw.get("location") or "").strip() or None,
            date_posted=_safe_date(raw.get("datePosted")),
            job_type=(raw.get("jobType") or "").strip() or None,
            is_remote=_parse_bool(raw.get("isRemote")),
            salary_min=_parse_float(raw.get("minAmount")),
            salary_max=_parse_float(raw.get("maxAmount")),
            currency=(raw.get("currency") or "").strip() or None,
            description=(raw.get("description") or "").strip() or None,
            normalized_hash=normalized_hash,
            matched_title=matched_title,
            matched_objective=matched_objective,
            raw_payload_json=raw,
        )
        db.add(job)
        created_jobs.append(job)
        per_site_counts[site_key] += 1

        metrics = company_metrics[key]
        metrics["jobs"] = int(metrics["jobs"]) + 1
        if job.location:
            metrics["locations"].append(job.location)
        metrics["sources"].append(source.display_name)
        if job.description:
            metrics["descriptions"].append(job.description)
        if job.date_posted:
            try:
                days_old = max((date.today() - date.fromisoformat(job.date_posted)).days, 0)
                metrics["days_active"].append(days_old)
            except Exception:
                pass
        if matched_signals:
            metrics["signals"].extend(matched_signals)
        if matched_title:
            metrics["title_matches"].append(title)

        company.open_roles = int(company.open_roles or 0) + 1
        company.source = source.display_name if not company.source else company.source
        company.revengineer_fit = fit if fit == "high" or company.revengineer_fit != "high" else company.revengineer_fit
        company.priority = priority if priority == "high" or company.priority != "high" else company.priority
        if matched_title and not company.title_match:
            company.title_match = title
        if matched_signals and not company.objective_signal:
            company.objective_signal = ", ".join(dict.fromkeys(matched_signals))
        if job.description and not company.description:
            company.description = job.description
        if not company.location and job.location:
            company.location = job.location
        if job_url and job_url not in company.web_sources:
            company.web_sources = [*company.web_sources, job_url]

    company_ids: list[str] = []
    matched_company_count = 0
    for key, company in companies_by_key.items():
        metrics = company_metrics[key]
        company_ids.append(company.id)
        if metrics["locations"]:
            company.location = Counter(metrics["locations"]).most_common(1)[0][0]
        if metrics["sources"]:
            company.source = " + ".join(sorted(dict.fromkeys(metrics["sources"])))
        if metrics["descriptions"] and not company.description:
            company.description = str(metrics["descriptions"][0])[:1200]
        if metrics["days_active"]:
            company.days_active = min(metrics["days_active"])
        signals = list(dict.fromkeys(str(item) for item in metrics["signals"]))
        title_matches = list(dict.fromkeys(str(item) for item in metrics["title_matches"]))
        fit, priority = _classify_fit(bool(title_matches), bool(signals))
        company.revengineer_fit = fit
        company.priority = priority
        company.title_match = company.title_match or (title_matches[0] if title_matches else None)
        company.objective_signal = company.objective_signal or (
            ", ".join(signals[:3]) if signals else None
        )
        company.web_evidence = company.web_evidence or (
            company.description[:400] if company.description else company.objective_signal
        )
        if signals:
            matched_company_count += 1

        db.add(
            models.CompanySignal(
                id=make_id("sig"),
                company_id=company.id,
                campaign_run_id=run.id,
                objective_score=min(10, 4 + len(signals) * 2 + (2 if title_matches else 0)),
                objective_classification=(
                    "likely" if signals else "possible" if title_matches else "unlikely"
                ),
                matched_signals=signals,
                evidence_snippet=company.web_evidence,
            )
        )

    for shortlist in shortlist_rows:
        key = _company_key(shortlist.get("companyName", ""))
        company = companies_by_key.get(key)
        if company is None:
            continue
        classification = (shortlist.get("classification") or "").strip().lower()
        score_raw = (shortlist.get("score") or "0").strip()
        try:
            objective_score = int(float(score_raw))
        except Exception:
            objective_score = 0
        matched_signals = [
            item.strip()
            for item in (shortlist.get("matchedSignals") or "").split(";")
            if item.strip()
        ]
        evidence = (shortlist.get("webEvidence") or shortlist.get("evidence") or "").strip()

        if classification == "likely":
            company.revengineer_fit = "high"
            company.priority = "high"
            matched_company_count += 1
        elif classification == "possible" and company.revengineer_fit == "low":
            company.revengineer_fit = "medium"
            company.priority = "medium"
        if evidence:
            company.web_evidence = evidence
        if shortlist.get("sampleTitle"):
            company.title_match = company.title_match or shortlist.get("sampleTitle")
        if matched_signals:
            company.objective_signal = ", ".join(matched_signals[:3])
        db.add(
            models.CompanySignal(
                id=make_id("sig"),
                company_id=company.id,
                campaign_run_id=run.id,
                objective_score=max(objective_score, 1),
                objective_classification=classification or "possible",
                matched_signals=matched_signals,
                evidence_snippet=evidence or company.web_evidence,
            )
        )

    artifact_records: list[models.Artifact] = []
    artifact_paths: list[Path] = []
    for output in output_files:
        copied = _copy_artifact(output.path, run.id)
        artifact_paths.append(copied)

    if company_ids:
        artifact_paths.append(_generate_shortlist_artifact(db, company_ids, run.id))

    summary_report = models.Report(
        id=make_id("rep"),
        campaign_run_id=run.id,
        name=f"{campaign.name} Output Summary",
        report_type="campaign_export",
        status="ready",
        focus="Campaign output",
        metric=f"{len(created_jobs)} jobs across {len(company_ids)} companies",
        summary=(
            f"Imported {len(created_jobs)} jobs from "
            f"{'fresh live scraper outputs' if ingestion_mode == 'fresh_live' else 'existing scraper outputs'} "
            f"and mapped them into {len(company_ids)} companies for this run."
        ),
        generated_at=datetime.utcnow(),
    )
    db.add(summary_report)
    db.flush()

    for artifact_path in artifact_paths:
        kind = artifact_path.suffix.lower().lstrip(".") or "file"
        mime_type = "text/csv" if kind == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        artifact = models.Artifact(
            id=make_id("art"),
            report_id=summary_report.id,
            kind=kind,
            file_name=artifact_path.name,
            mime_type=mime_type,
            file_path=str(artifact_path),
        )
        db.add(artifact)
        artifact_records.append(artifact)

    if company_ids:
        shortlist_report = models.Report(
            id=make_id("rep"),
            campaign_run_id=run.id,
            name="US Target Companies Shortlist",
            report_type="company_shortlist",
            status="ready",
            focus="Company targeting",
            metric=f"{matched_company_count} companies with objective signals",
            summary=(
                "Companies aggregated from real scraper outputs whose descriptions or titles match "
                "the active objective filters."
            ),
            generated_at=datetime.utcnow(),
        )
        db.add(shortlist_report)
        db.flush()

        shortlist_artifact_path = settings.artifacts_dir / f"{run.id}-company-shortlist.csv"
        db.add(
            models.Artifact(
                id=make_id("art"),
                report_id=shortlist_report.id,
                kind="csv",
                file_name=shortlist_artifact_path.name,
                mime_type="text/csv",
                file_path=str(shortlist_artifact_path),
            )
        )

    for site_key, jobs_found in per_site_counts.items():
        health = ensure_source_health(db, site_key)
        credential = ensure_source_credential(db, site_key)
        health.status = "ready"
        health.last_success_at = datetime.utcnow()
        health.last_error_at = None
        health.last_error_message = None
        health.last_run_jobs_found = jobs_found
        health.avg_results_7d = float(jobs_found)
        health.success_rate_7d = 100.0
        credential.working_status = "working"
        credential.credential_present = credential.credential_present or (
            (credential.needs_api_key and bool(settings.scrappa_token))
            or (credential.needs_proxy and bool(settings.jobs_bota_proxy or settings.jobs_proxy))
            or (not credential.needs_api_key and not credential.needs_proxy)
        )
        if credential.credential_present:
            credential.credential_verified_at = datetime.utcnow()
        record_source_evidence(
            db,
            site_key=site_key,
            run_id=run.id,
            evidence_type=source_evidence_type or ingestion_mode,
            jobs_found=jobs_found,
            succeeded=jobs_found > 0,
            query_signature=campaign.role_query,
            country=campaign.country,
            location=campaign.location,
            details={
                "ingestionMode": ingestion_mode,
                "matchedTitleFiltering": restrict_to_campaign_title_match,
            },
        )
        refresh_source_support(db, site_key, health=health, credential=credential)

    if ats_rows:
        ats_counts = Counter(_normalize_site_key(row.get("ats")) for row in ats_rows if row.get("ats"))
        for site_key, jobs_found in ats_counts.items():
            health = ensure_source_health(db, site_key)
            credential = ensure_source_credential(db, site_key)
            health.status = "ready"
            health.last_success_at = datetime.utcnow()
            health.last_run_jobs_found = jobs_found
            health.avg_results_7d = float(jobs_found)
            health.success_rate_7d = 100.0
            per_site_counts[site_key] += jobs_found
            credential.working_status = "working"
            credential.credential_present = credential.credential_present or bool(settings.scrappa_token)
            if credential.credential_present:
                credential.credential_verified_at = datetime.utcnow()
            record_source_evidence(
                db,
                site_key=site_key,
                run_id=run.id,
                evidence_type=source_evidence_type or ingestion_mode,
                jobs_found=jobs_found,
                succeeded=jobs_found > 0,
                query_signature=campaign.role_query,
                country=campaign.country,
                location=campaign.location,
                details={"ingestionMode": ingestion_mode, "atsImport": True},
            )
            refresh_source_support(db, site_key, health=health, credential=credential)

    source_summary_map = {item.get("siteKey"): item for item in source_summary}
    for site_key, jobs_found in per_site_counts.items():
        summary_item = source_summary_map.get(site_key)
        if summary_item is None:
            source_summary_map[site_key] = {
                "siteKey": site_key,
                "status": "completed" if jobs_found else "failed",
                "jobsFound": jobs_found,
                "durationMs": None,
                "error": None if jobs_found else "No rows imported from source output.",
            }
        else:
            summary_item["jobsFound"] = jobs_found
            summary_item["status"] = "completed" if jobs_found else summary_item.get("status", "failed")
            if jobs_found:
                summary_item["error"] = None

    if mark_empty_results_as_failed:
        for summary_item in source_summary_map.values():
            if int(summary_item.get("jobsFound") or 0) > 0:
                continue
            summary_item["status"] = "failed"
            summary_item["error"] = summary_item.get("error") or "No rows matched the current campaign title filters."

    return {
        "job_count": len(created_jobs),
        "company_count": len(company_ids),
        "matched_company_count": matched_company_count,
        "source_summary": list(source_summary_map.values()),
    }
